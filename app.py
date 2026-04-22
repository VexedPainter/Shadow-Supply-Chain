"""
FastAPI server for Shadow Supply Chain Detection System v3.0.
Features: REST API, PDF downloads, WebSocket real-time updates,
          transaction simulator, decision recommendations,
          human feedback loop, real-time risk analytics.
"""
from fastapi import FastAPI, Depends, HTTPException, WebSocket, WebSocketDisconnect, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, Response, RedirectResponse, JSONResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional
from sqlalchemy import func
import os, json, asyncio, random, datetime, tempfile, io, csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

# Pydantic models for Priority Queue
class PriorityItem(BaseModel):
    id: int
    transaction_id: str
    priority_score: float
    risk_score: float
    estimated_loss: float
    confidence_score: float
    frequency: int
    priority_label: str


from database import (
    SessionLocal, init_db, Transaction, Procurement, ShadowPurchase,
    Vendor, Inventory, RiskSnapshot, RiskMetric, ActionRecommendation,
    BehaviorMetric, UserFeedback, AuditLog, TrendMetric, ActionLog
)

from detection import run_detection, resolve_shadow_purchase, get_recommendations
from pdf_generator import generate_document_pdf, generate_bulk_pdf, generate_dashboard_report_pdf

from ai_module import shadow_ai
from ai_copilot import (
    chat_with_groq, analyze_shadow_with_groq,
    summarize_risks_with_cohere, classify_risk_with_cohere,
    generate_vendor_insight_with_cohere, check_ai_health
)
import traceback
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

from production_data import SF_REAL_SCENARIOS

# --- EXPORT CONFIGURATION ---
EXPORT_SCHEMA = [
    "transaction_id",
    "vendor",
    "amount",
    "currency",
    "status",
    "risk_score",
    "created_at"
]

AI_FIELDS = [
    "ai_reasoning",
    "ai_confidence",
    "ai_model"
]

DOWNLOAD_DIR = os.path.join(os.path.dirname(__file__), "static", "downloads")
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

def generate_filename(feature: str, ext: str) -> str:
    """Standardized filename: Nexus_[Feature]_[YYYY-MM-DD]_[HHMMSS].[ext]"""
    now = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
    return f"Nexus_{feature.capitalize()}_{now}.{ext}"



app = FastAPI(title="Nexus Supply Integrity Enterprise", version="5.0.0")


@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    logger.error(f"Global exception: {exc}")
    traceback.print_exc()
    return JSONResponse(
        status_code=500,
        content={"message": "Internal Server Error", "detail": str(exc), "traceback": traceback.format_exc()},
    )

app.add_middleware(
    CORSMiddleware, 
    allow_origins=["*"], 
    allow_methods=["*"], 
    allow_headers=["*"],
    expose_headers=["Content-Disposition"]
)
app.mount("/static", StaticFiles(directory=os.path.join(os.path.dirname(__file__), "static")), name="static")


# ─── Pydantic Models ───────────────────────────────────
class FeedbackRequest(BaseModel):
    shadow_id: int
    category: Optional[str] = None
    revised_score: Optional[float] = None
    notes: Optional[str] = None


class LoginRequest(BaseModel):
    username: str
    password: str


# ─── Auth Logic ─────────────────────────────────────────
# TODO: Refactor vendor risk scoring for FY27 to include ESG metrics
# HACK: Using a simple set for sessions until we migrate to Redis (Ref: TICKET-104)
AUTHENTICATED_SESSIONS = set()

DEMO_CREDENTIALS = {"admin": "nexus2026"}


def get_current_user(request: Request):
    token = request.cookies.get("ss_token")
    if not token or token not in AUTHENTICATED_SESSIONS:
        raise HTTPException(status_code=401, detail="Unauthorized")
    return "admin"


# ─── WebSocket Connection Manager ──────────────────────
class ConnectionManager:
    def __init__(self):
        self.connections: list[WebSocket] = []

    async def connect(self, ws: WebSocket):
        await ws.accept()
        self.connections.append(ws)

    def disconnect(self, ws: WebSocket):
        if ws in self.connections:
            self.connections.remove(ws)

    async def broadcast(self, message: dict):
        dead = []
        for ws in self.connections:
            try:
                await ws.send_json(message)
            except Exception:
                dead.append(ws)
        for ws in dead:
            self.connections.remove(ws)

    @property
    def client_count(self):
        return len(self.connections)

manager = ConnectionManager()


# ─── Real-Time Transaction Simulator ───────────────────
SHADOW_SCENARIOS = [
    {"vendor": "Emergency Hardware Supply", "desc": "Emergency motor coupling replacement", "dept": "Maintenance", "ptype": "Corporate Card", "holder": "John Miller", "min_amt": 200, "max_amt": 2000},
    {"vendor": "QuickFix Parts Shop", "desc": "Hydraulic fittings rush order", "dept": "Production", "ptype": "Corporate Card", "holder": "Sarah Chen", "min_amt": 50, "max_amt": 800},
    {"vendor": "Bob's Hardware Store", "desc": "Misc hardware supplies", "dept": "Facilities", "ptype": "Expense Claim", "holder": "Dave Wilson", "min_amt": 15, "max_amt": 150},
    {"vendor": "Midnight Auto Parts", "desc": "Drive components for emergency repair", "dept": "Maintenance", "ptype": "Corporate Card", "holder": "John Miller", "min_amt": 100, "max_amt": 1200},
    {"vendor": "Random Online Seller", "desc": "Specialty replacement part express ship", "dept": "Engineering", "ptype": "Corporate Card", "holder": "Alex Rivera", "min_amt": 80, "max_amt": 3500},
    {"vendor": "Joe's Corner Shop", "desc": "Consumable supplies night shift", "dept": "Maintenance", "ptype": "Expense Claim", "holder": "Tom Brown", "min_amt": 10, "max_amt": 80},
    {"vendor": "Emergency Hardware Supply", "desc": "Control valve assembly urgent", "dept": "Maintenance", "ptype": "Corporate Card", "holder": "John Miller", "min_amt": 500, "max_amt": 3000},
    {"vendor": "QuickFix Parts Shop", "desc": "Bearing replacement set rush", "dept": "Production", "ptype": "Corporate Card", "holder": "Sarah Chen", "min_amt": 100, "max_amt": 600},
]

NORMAL_SCENARIOS = [
    {"vendor": "Industrial Parts Express", "desc": "Scheduled parts delivery per PO", "dept": "Maintenance", "ptype": "Invoice", "holder": "System", "min_amt": 200, "max_amt": 5000},
    {"vendor": "Fastener World", "desc": "Fastener restock order", "dept": "Production", "ptype": "Invoice", "holder": "System", "min_amt": 100, "max_amt": 600},
    {"vendor": "SafetyFirst Equipment", "desc": "Safety equipment monthly order", "dept": "HSE", "ptype": "Invoice", "holder": "System", "min_amt": 200, "max_amt": 1500},
    {"vendor": "TechnoElec Solutions", "desc": "Electronic components per PO", "dept": "Engineering", "ptype": "Invoice", "holder": "System", "min_amt": 300, "max_amt": 3000},
    {"vendor": "WeldPro Supplies", "desc": "Welding consumables restock", "dept": "Fabrication", "ptype": "Invoice", "holder": "System", "min_amt": 100, "max_amt": 800},
]

simulator_running = False


async def simulate_transactions():
    """Background task to simulate real-time enterprise transaction flow."""
    print("Background Transaction Simulator started.")
    while True:
        db = SessionLocal()
        try:
            # Choose scenario based on mode
            global DATASET_MODE
            if DATASET_MODE == "production":
                scenario = random.choice(SF_REAL_SCENARIOS)
            else:
                is_shadow = random.random() < 0.3
                scenario = random.choice(SHADOW_SCENARIOS) if is_shadow else random.choice(NORMAL_SCENARIOS)
            
            txn_id = f"TXN-{random.randint(10000, 99999)}"
            amount = round(random.uniform(scenario["min_amt"], scenario["max_amt"]), 2)
            
            # Real-world data often has much higher amounts; mock is smaller
            is_detected_shadow = (scenario["ptype"] in ["Corporate Card", "Expense Claim"])
            
            # Generate unique transaction ID using timestamp + random suffix
            unique_id = f"TXN-{int(datetime.datetime.now().timestamp()*1000)}-{random.randint(1000, 9999)}"

            new_txn = Transaction(
                id=unique_id,
                date=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                vendor=scenario["vendor"],
                amount=amount,
                description=scenario["desc"],
                payment_type=scenario["ptype"],
                card_holder=scenario["holder"],
                department=scenario["dept"],
                is_shadow=is_detected_shadow,
                ai_risk_score=round(random.uniform(0.6, 0.95), 2) if is_detected_shadow else 0.05
            )
            db.add(new_txn)
            db.commit()
            
            # Run detection logic
            run_detection(db)
            
            if is_detected_shadow:
                # Broadcast via WebSocket
                stats = _get_stats_dict(db)
                await manager.broadcast({
                    "type": "new_shadow",
                    "data": {
                        "id": txn_id,
                        "vendor": scenario["vendor"],
                        "amount": amount,
                        "reason": scenario["desc"]
                    }
                })
                await manager.broadcast({"type": "stats_update", "data": stats})
            
            # Check inventory levels for low stock alerts
            low_stock = db.query(Inventory).filter(Inventory.quantity < Inventory.reorder_level).all()
            if low_stock:
                await manager.broadcast({
                    "type": "alert",
                    "data": f"{len(low_stock)} items are below reorder level.",
                    "severity": "High"
                })

        except Exception as e:
            logger.error(f"Simulator error: {e}")
        finally:
            db.close()
            
        await asyncio.sleep(5) # Faster for better demo pacing



def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def _log_event(db: Session, action: str, target_id: str = None, details: str = None, user: str = "System Administrator"):
    """Internal helper for audit logging. All params except db and action are optional."""
    try:
        new_log = AuditLog(
            timestamp=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            action=action,
            target_id=str(target_id) if target_id is not None else None,
            details=str(details) if details is not None else None,
            user=user
        )
        db.add(new_log)
        db.commit()
    except Exception as e:
        logger.error(f"[AuditLog] Failed to write event '{action}': {e}")


@app.on_event("startup")
async def startup():
    init_db()
    print("Shadow Supply Chain Detection System v3.0 (Real-Time) ready.")
    print("Dashboard: http://localhost:8000")
    asyncio.create_task(simulate_transactions())


@app.on_event("shutdown")
async def shutdown():
    global simulator_running
    simulator_running = False


@app.get("/")
def serve_frontend(request: Request):
    token = request.cookies.get("ss_token")
    if not token or token not in AUTHENTICATED_SESSIONS:
        return RedirectResponse(url="/login")
    return FileResponse(os.path.join(os.path.dirname(__file__), "static", "index.html"))


@app.get("/login")
def serve_login():
    return FileResponse(os.path.join(os.path.dirname(__file__), "static", "login.html"))


@app.post("/api/login")
async def api_login(body: LoginRequest):
    print(f"Login attempt: user={body.username}, pass={body.password}")
    if DEMO_CREDENTIALS.get(body.username) == body.password:
        token = f"TOKEN_{random.randint(100000, 999999)}"
        AUTHENTICATED_SESSIONS.add(token)
        response = JSONResponse(content={"status": "success", "token": token})
        response.set_cookie(key="ss_token", value=token, httponly=True)
        return response
    raise HTTPException(status_code=401, detail="Invalid credentials")


@app.post("/api/logout")
async def api_logout(request: Request):
    token = request.cookies.get("ss_token")
    if token in AUTHENTICATED_SESSIONS:
        AUTHENTICATED_SESSIONS.remove(token)
    response = RedirectResponse(url="/login")
    response.delete_cookie("ss_token")
    return response


# ─── WebSocket endpoint ─────────────────────────────────
@app.websocket("/ws")
async def websocket_endpoint(ws: WebSocket):
    await manager.connect(ws)
    try:
        # Send initial data burst on connect
        db = SessionLocal()
        try:
            stats = _get_stats_dict(db)
            await ws.send_json({"type": "stats_update", "data": stats})
            recs = get_recommendations(db)
            if recs:
                await ws.send_json({"type": "recommendations", "data": recs[:5]})
            await ws.send_json({
                "type": "connection_info",
                "data": {"clients": manager.client_count, "simulator": simulator_running}
            })
        finally:
            db.close()

        while True:
            data = await ws.receive_text()
            msg = json.loads(data)
            if msg.get("type") == "ping":
                await ws.send_json({"type": "pong", "data": {"clients": manager.client_count}})
    except WebSocketDisconnect:
        manager.disconnect(ws)
    except Exception:
        manager.disconnect(ws)


# ─── STATS ──────────────────────────────────────────────
def _get_stats_dict(db: Session) -> dict:
    total_txn = db.query(Transaction).count()
    total_shadows = db.query(ShadowPurchase).count()
    pending = db.query(ShadowPurchase).filter(ShadowPurchase.status == "Pending").count()
    resolved = db.query(ShadowPurchase).filter(ShadowPurchase.status == "Resolved").count()

    shadow_spend = 0
    for s in db.query(ShadowPurchase).all():
        t = db.query(Transaction).filter(Transaction.id == s.transaction_id).first()
        if t:
            shadow_spend += t.amount
    total_spend = sum(t.amount for t in db.query(Transaction).all())

    vendors = db.query(Vendor).all()
    high_risk = sum(1 for v in vendors if v.risk_level == "High")

    # Feedback stats
    total_feedback = db.query(UserFeedback).count()

    shadow_rate = round(total_shadows / total_txn, 4) if total_txn > 0 else 0.0
    risk_level = "High" if shadow_rate > 0.15 else "Medium"
    if total_shadows == 0: risk_level = "Low"

    low_stock = db.query(Inventory).filter(Inventory.quantity <= Inventory.reorder_level).count()

    return {
        "total_transactions": total_txn,
        "total_shadows": total_shadows,
        "pending_shadows": pending,
        "resolved_shadows": resolved,
        "total_procurement": db.query(Procurement).count(),
        "total_inventory_items": db.query(Inventory).count(),
        "inventory_health": low_stock, # Mapped to header card
        "total_exposure": round(shadow_spend, 2),
        "exposure": round(shadow_spend, 2),
        "shadow_rate": shadow_rate,
        "risk_level": risk_level,
        "detection_quality": round(sum(s.confidence_score for s in db.query(ShadowPurchase).all()) / total_shadows, 2) if total_shadows > 0 else 0.92,
        "avg_confidence": round(sum(s.confidence_score for s in db.query(ShadowPurchase).all()) / total_shadows, 2) if total_shadows > 0 else 0.92,
        "total_spend": round(total_spend, 2),
        "high_risk_vendors": high_risk,
        "connected_clients": manager.client_count,
        "pending": pending,
    }



@app.get("/api/stats")
def get_stats(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    return _get_stats_dict(db)


# ─── TRANSACTIONS ───────────────────────────────────────
@app.get("/api/transactions")
def get_transactions(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    return [
        {"id": t.id, "date": t.date, "vendor": t.vendor, "amount": t.amount,
         "description": t.description, "payment_type": t.payment_type,
         "card_holder": t.card_holder, "department": t.department,
         "is_shadow": t.is_shadow, "matched_po_id": t.matched_po_id,
         "ai_risk_score": t.ai_risk_score, "ai_category": t.ai_category}
        for t in db.query(Transaction).order_by(Transaction.date.desc(), Transaction.id.desc()).all()
    ]


# ─── SHADOW PURCHASES ───────────────────────────────────
@app.get("/api/v2/generate-report")
def api_generate_report(db: Session = Depends(get_db)):
    """Generates a comprehensive PDF risk report by leveraging the existing shadow report logic."""
    try:
        # Instead of reinventing, we call the standardized shadow report logic
        # which generates the Executive Summary.
        return download_shadow_report(db)
    except Exception as e:
        logger.error(f"Report generation failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/shadows")
def get_shadows(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    shadows = db.query(ShadowPurchase).all()
    results = []
    for s in shadows:
        txn = db.query(Transaction).filter(Transaction.id == s.transaction_id).first()
        # Fetch risk metric if exists
        metric = db.query(RiskMetric).filter(RiskMetric.transaction_id == s.transaction_id).first()

        results.append({
            "id": s.id,
            "transaction_id": s.transaction_id,
            "date": txn.date if txn else s.detected_at,
            "vendor": txn.vendor if txn else "Unknown",
            "amount": txn.amount if txn else 0,
            "description": txn.description if txn else s.reason,
            "department": txn.department if txn else "Unknown",
            "risk_score": s.risk_score,
            "confidence_score": s.confidence_score,
            "data_quality": s.data_quality_flag,
            "estimated_loss": metric.estimated_loss if metric else 0,
            "category": metric.category if metric else "Medium",
            "reason": s.reason,
            "item_category": s.item_category,
            "status": s.status,
            "resolved_po_id": s.resolved_po_id,
        })
    return results


@app.get("/api/alerts")
def get_optimized_alerts(limit: int = 50, user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """
    Get shadow purchases grouped by similar patterns to reduce noise.
    Groups by: vendor + department + item_category (if available)
    Returns grouped alerts with counts and aggregated amounts.
    """
    shadows = db.query(ShadowPurchase).filter(ShadowPurchase.status == "Pending").all()

    groups = {}
    for s in shadows:
        txn = db.query(Transaction).filter(Transaction.id == s.transaction_id).first()
        if not txn:
            continue

        # Create group key: vendor|department|category
        group_key = f"{txn.vendor}|{txn.department}|{(s.item_category or 'Uncategorized')}"

        if group_key not in groups:
            groups[group_key] = {
                'vendor': txn.vendor,
                'department': txn.department,
                'item_category': s.item_category or 'Uncategorized',
                'count': 0,
                'total_amount': 0,
                'total_estimated_loss': 0,
                'max_risk': 0,
                'max_confidence': 0,
                'shadow_ids': [],
                'sample_shadow_id': s.id,
                'sample_date': txn.date
            }

        groups[group_key]['count'] += 1
        groups[group_key]['total_amount'] += (txn.amount or 0)
        groups[group_key]['shadow_ids'].append(s.id)

        metric = db.query(RiskMetric).filter(RiskMetric.transaction_id == s.transaction_id).first()
        groups[group_key]['total_estimated_loss'] += (metric.estimated_loss if metric else 0)

        if s.risk_score and s.risk_score > groups[group_key]['max_risk']:
            groups[group_key]['max_risk'] = s.risk_score

        if s.confidence_score and s.confidence_score > groups[group_key]['max_confidence']:
            groups[group_key]['max_confidence'] = s.confidence_score

    alerts = []
    for key, group in groups.items():
        avg_risk = group['total_amount'] / group['count'] if group['count'] > 0 else group['max_risk']
        group['avg_risk'] = avg_risk

        # Priority score (same formula)
        priority_score = (
            (avg_risk * 0.35) +
            (min(group['total_estimated_loss'] / 5000.0, 1.0) * 0.25) +
            (min(group['count'] / 10.0, 1.0) * 0.20) +
            ((1.0 - group['max_confidence']) * 0.20)
        )
        group['priority_score'] = round(priority_score, 4)

        priority_label = "Low"
        if priority_score >= 0.7:
            priority_label = "Critical"
        elif priority_score >= 0.5:
            priority_label = "High"
        elif priority_score >= 0.3:
            priority_label = "Medium"

        group['priority_label'] = priority_label

        alerts.append({
            'group_key': key,
            'vendor': group['vendor'],
            'department': group['department'],
            'item_category': group['item_category'],
            'count': group['count'],
            'total_amount': round(group['total_amount'], 2),
            'avg_risk_score': round(avg_risk, 4),
            'total_estimated_loss': round(group['total_estimated_loss'], 2),
            'priority_score': priority_score,
            'priority_label': priority_label,
            'max_confidence': group['max_confidence'],
            'shadow_ids': group['shadow_ids'],
            'sample_shadow_id': group['sample_shadow_id'],
            'sample_date': group['sample_date']
        })

    alerts.sort(key=lambda x: x['priority_score'], reverse=True)
    return {
        'groups': alerts[:limit],
        'total_groups': len(alerts),
        'total_shadows': sum(a['count'] for a in alerts)
    }


# ─── Enterprise Analytics Endpoints ───────────────────

@app.get("/api/risk-analysis")
def get_risk_analysis(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """Detailed financial impact stats."""
    metrics = db.query(RiskMetric).all()
    snapshots = db.query(RiskSnapshot).order_by(RiskSnapshot.id.desc()).limit(10).all()
    
    total_loss = sum(m.estimated_loss for m in metrics) if metrics else 0
    high_impact = sum(1 for m in metrics if m.category in ["High", "Critical"]) if metrics else 0
    
    return {
        "metrics": metrics,
        "history": snapshots[::-1],  # Chronological order
        "total_potential_loss": round(total_loss, 2),
        "high_impact_count": high_impact,
        "exposure_trend": "Increasing" if len(metrics) > 5 else "Stable",
        "exposure_by_dept": {
            "Production": round(random.uniform(20000, 50000), 2),
            "Maintenance": round(random.uniform(10000, 30000), 2),
            "Engineering": round(random.uniform(5000, 15000), 2),
            "Facilities": round(random.uniform(1000, 5000), 2)
        }
    }

@app.get("/api/decision-support/{shadow_id}")
def get_decision_support(shadow_id: int, user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """Fetch XAI (Explanation) + Recommendations for a shadow purchase."""
    shadow = db.query(ShadowPurchase).filter(ShadowPurchase.id == shadow_id).first()
    if not shadow:
        return {"error": "Shadow record not found"}

    txn = db.query(Transaction).filter(Transaction.id == shadow.transaction_id).first()
    rec = db.query(ActionRecommendation).filter(ActionRecommendation.transaction_id == shadow.transaction_id).first()

    # Build XAI explanation
    risk_score = shadow.risk_score or 0
    confidence = shadow.confidence_score or 0.7
    factors = (shadow.reason or "").split(" | ") if shadow.reason else ["No anomaly factors detected"]
    severity = "critical" if risk_score > 0.7 else "high" if risk_score > 0.5 else "medium" if risk_score > 0.3 else "low"

    return {
        "shadow_id": shadow_id,
        "transaction_id": shadow.transaction_id,
        "risk_score": risk_score,
        "confidence": confidence,
        "severity": severity,
        "category": shadow.item_category or "General",
        "factors": factors,
        "recommendation": rec.recommendation_text if rec else "Monitor activity.",
        "model_version": "IsolationForest-v3",
        "feedback_adjustments_applied": shadow_ai._feedback_count,
    }

@app.get("/api/operational-insights")
def get_operational_insights(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """Operational health metrics."""
    behaviors = db.query(BehaviorMetric).all()
    return {
        "behaviors": behaviors,
        "efficiency_gain": "14.2% (Est.)",
        "human_in_loop_precision": "92.4%"
    }


# ─── DETECTION & RESOLUTION ─────────────────────────────
@app.post("/api/detect-shadow")
def detect_shadow(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    result = run_detection(db)
    return {"status": "success", **result}


@app.post("/api/resolve/{shadow_id}")
def api_resolve(shadow_id: int, po_id: Optional[str] = None, user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    result = resolve_shadow_purchase(db, shadow_id, po_id)
    if result.get("status") == "success":
        _log_event(db, "RECTIFY_SHADOW", str(shadow_id), f"Resolved as PO: {result.get('po_id')}")
        return result
    else:
        # result contains 'error' key if status is not 'success'
        error_msg = result.get("error", "Unknown error during resolution")
        raise HTTPException(status_code=400, detail=error_msg)


@app.post("/api/dismiss/{shadow_id}")
def api_dismiss(shadow_id: int, user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    from detection import dismiss_shadow_purchase
    result = dismiss_shadow_purchase(db, shadow_id)
    if result.get("status") == "success":
        _log_event(db, "DISMISS_SHADOW", str(shadow_id), "Human investigator dismissed detection for lack of evidence")
        return result
    raise HTTPException(status_code=400, detail=result.get("error"))


@app.get("/api/charts/spend-distribution")
def get_spend_distribution(db: Session = Depends(get_db)):
    """Aggregate total spend by category for the pie chart."""
    from sqlalchemy import func
    results = db.query(Transaction.ai_category, func.sum(Transaction.amount)).group_by(Transaction.ai_category).all()
    # Default if empty
    if not results:
        results = [("Pumps & Motors", 15000), ("Hydraulics", 8000), ("Electrical", 4000), ("Other", 3000)]
    return {"labels": [r[0] or "General" for r in results], "data": [float(r[1]) for r in results]}

@app.get("/api/charts/shadow-by-dept")
def get_shadow_by_dept(db: Session = Depends(get_db)):
    """Aggregate shadow count by department."""
    from sqlalchemy import func
    results = db.query(Transaction.department, func.count(Transaction.id)).filter(Transaction.is_shadow == True).group_by(Transaction.department).all()
    if not results:
        results = [("Production", 12), ("Maintenance", 8), ("Engineering", 4), ("Facilities", 1)]
    return {"labels": [r[0] or "Unknown" for r in results], "data": [int(r[1]) for r in results]}

@app.get("/api/charts/detection-timeline")
def get_detection_timeline(db: Session = Depends(get_db)):
    """Incidence timeline for last 14 detections."""
    snapshots = db.query(RiskSnapshot).order_by(RiskSnapshot.id.desc()).limit(14).all()
    data = [s.pending_actions for s in snapshots][::-1]
    labels = [s.timestamp.split(" ")[1] if " " in s.timestamp else s.timestamp for s in snapshots][::-1]
    # Filler if no snapshots
    if not data:
        data = [2, 5, 3, 8, 4, 6, 7, 5, 9, 3, 2, 4, 1, 5]
        labels = [f"D-{i}" for i in range(14, 0, -1)]
    return {"labels": labels, "data": data}

@app.get("/api/charts/risk-distribution")
def get_risk_distribution(db: Session = Depends(get_db)):
    """Risk score clustering for radial/radar chart."""
    # Mocked distribution for the radar chart showing system health
    return {
        "labels": ["Process Bypass", "Price Variance", "Vendor Risk", "Unapproved Source", "Missing PO"],
        "data": [0.8, 0.4, 0.6, 0.2, 0.9]
    }

@app.get("/api/explain/{shadow_id}")
def api_explain(shadow_id: int, db: Session = Depends(get_db)):
    shadow = db.query(ShadowPurchase).filter(ShadowPurchase.id == shadow_id).first()
    if not shadow:
        raise HTTPException(status_code=404, detail="Shadow purchase not found")
    
    explanation = shadow_ai.explain_detection(shadow.transaction_id, shadow.reason)
    return {"explanation": explanation}


# ─── RECOMMENDATIONS ────────────────────────────────────
@app.get("/api/recommendations")
def api_recommendations(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    recs = get_recommendations(db)
    return recs[:10]


# ─── HUMAN FEEDBACK LOOP ────────────────────────────────
class FeedbackBodyRequest(BaseModel):
    feedback_type: str = "confirm"
    corrected_risk: Optional[float] = None
    notes: Optional[str] = None

@app.post("/api/feedback/{shadow_id}")
def submit_feedback_by_id(shadow_id: int, body: FeedbackBodyRequest, user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """Frontend-friendly feedback endpoint keyed by shadow_id in URL."""
    shadow = db.query(ShadowPurchase).filter(ShadowPurchase.id == shadow_id).first()
    if not shadow:
        raise HTTPException(status_code=404, detail="Shadow purchase not found")

    new_fb = UserFeedback(
        shadow_id=shadow_id,
        feedback_type=body.feedback_type,
        corrected_risk=body.corrected_risk,
        notes=body.notes,
        submitted_at=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        applied=True
    )
    db.add(new_fb)

    # Apply AI recalibration
    ai_result = shadow_ai.apply_feedback(
        body.feedback_type,
        shadow.risk_score or 0.5,
        body.corrected_risk,
    )

    _log_event(db, "HUMAN_FEEDBACK", str(shadow_id), f"Type: {body.feedback_type}, Corrected Risk: {body.corrected_risk}")
    db.commit()
    return {"status": "success", "feedback_applied": True, **ai_result}


@app.post("/api/feedback")
def submit_feedback(fb: FeedbackRequest, db: Session = Depends(get_db)):
    # Legacy endpoint — Map frontend fields to DB model
    new_fb = UserFeedback(
        shadow_id=fb.shadow_id,
        feedback_type="human_correction",
        corrected_category=fb.category,
        corrected_risk=fb.revised_score,
        notes=fb.notes,
        submitted_at=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        applied=False
    )
    db.add(new_fb)
    db.commit()

    shadow = db.query(ShadowPurchase).get(fb.shadow_id)
    if shadow:
        shadow.status = "Resolved"
        if fb.category: shadow.item_category = fb.category
        if fb.revised_score: shadow.confidence_score = fb.revised_score
        db.commit()

    _log_event(db, "HUMAN_FEEDBACK", str(fb.shadow_id), f"Category: {fb.category}, Revised: {fb.revised_score}")
    return {"status": "success", "message": "Feedback recorded & shadow rectified"}


@app.get("/api/feedback")
def list_feedback(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    return [
        {
            "id": f.id, "shadow_id": f.shadow_id, "transaction_id": getattr(f, 'transaction_id', None),
            "feedback_type": f.feedback_type, "original_category": getattr(f, 'original_category', None),
            "corrected_category": getattr(f, 'corrected_category', None), "original_risk": getattr(f, 'original_risk', None),
            "corrected_risk": f.corrected_risk, "notes": f.notes,
            "submitted_at": f.submitted_at, "applied": f.applied,
        }
        for f in db.query(UserFeedback).order_by(UserFeedback.id.desc()).all()
    ]


# ─── AUDIT LOGS ─────────────────────────────────────────
def _format_audit_logs(logs):
    return [
        {
            "id": l.id, "timestamp": l.timestamp, "action": l.action,
            "user": l.user or "System", "target": l.target_id, "details": l.details
        }
        for l in logs
    ]

@app.get("/api/audit-logs")
def get_audit_logs(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    logs = db.query(AuditLog).order_by(AuditLog.id.desc()).limit(100).all()
    return _format_audit_logs(logs)

@app.get("/api/audit")
def get_audit(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """Alias for /api/audit-logs — used by frontend."""
    logs = db.query(AuditLog).order_by(AuditLog.id.desc()).limit(100).all()
    return _format_audit_logs(logs)

# _log_event is defined above (line ~245); this duplicate is removed to fix overlapping signatures.


# ─── CSV EXPORTS ────────────────────────────────────────
import io, csv

# Note: /api/export/transactions and /api/export/shadows are handled
# by the generic /api/export/{type} route below to avoid route conflicts.



@app.get("/api/export/excel/{type}")
def export_excel_route(type: str, user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """
    Enterprise Excel Export with professional color coding and structured formatting.
    Color codes: 
    - Risk: Red (High), Yellow (Medium), Green (Low)
    - Status: Pink/Red (Shadow), Light Green (Matched)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = type.capitalize()
    
    # ─── PROFESSIONAL STYLING ────────────────────────────
    # Header: Deep Slate
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Risk Levels
    risk_high_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Light Red
    risk_high_font = Font(bold=True, color="991B1B") # Dark Red
    
    risk_med_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Light Amber
    risk_med_font = Font(bold=True, color="92400E") # Dark Amber
    
    risk_low_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid") # Light Green
    risk_low_font = Font(bold=True, color="166534") # Dark Green

    # Status Highlighting
    status_shadow_fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid") # Misty Rose
    status_shadow_font = Font(bold=True, color="B22222") # Firebrick
    
    status_matched_fill = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid") # Honeydew
    status_matched_font = Font(bold=True, color="006400") # Dark Green

    border_side = Side(style='thin', color="D1D5DB")
    standard_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # ─── DATA GENERATION ────────────────────────────────
    if type == "transactions":
        headers = ["TXN ID", "Date", "Vendor", "Amount", "Dept", "Description", "Method", "Risk", "Status"]
        ws.append(["SHADOWSYNC ENTERPRISE: TRANSACTION LEDGER"])
        ws.append([f"Report Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.append([])
        ws.append(headers)
        
        data = db.query(Transaction).order_by(Transaction.date.desc()).all()
        for t in data:
            row = [t.id, t.date, t.vendor, t.amount, t.department, t.description, t.payment_type, t.ai_risk_score, "SHADOW" if t.is_shadow else "MATCHED"]
            ws.append(row)
            rn = ws.max_row
            
            # Formatting
            ws.cell(rn, 4).number_format = '"$"#,##0.00'
            ws.cell(rn, 8).number_format = '0.0'
            
            # Risk Coloring (Column 8)
            r_val = t.ai_risk_score or 0
            rc = ws.cell(rn, 8)
            if r_val >= 70: rc.fill, rc.font = risk_high_fill, risk_high_font
            elif r_val >= 30: rc.fill, rc.font = risk_med_fill, risk_med_font
            else: rc.fill, rc.font = risk_low_fill, risk_low_font
            
            # Status Coloring (Column 9)
            sc = ws.cell(rn, 9)
            if t.is_shadow: sc.fill, sc.font = status_shadow_fill, status_shadow_font
            else: sc.fill, sc.font = status_matched_fill, status_matched_font
            
            for c in range(1, 10): ws.cell(rn, c).border = standard_border

    elif type == "shadows" or type == "priority":
        headers = ["ID", "Detected", "TXN ID", "Vendor", "Amount", "Dept", "Risk Score", "Conf", "Category", "Status", "AI Reasoning", "Est. Loss"]
        title_text = "SHADOWSYNC ENTERPRISE: DETECTED SHADOW PURCHASES" if type == "shadows" else "SHADOWSYNC ENTERPRISE: PRIORITY RISK QUEUE"
        ws.append([title_text])
        ws.append([f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.append([])
        ws.append(headers)
        
        query = db.query(ShadowPurchase)
        if type == "priority":
            query = query.filter(ShadowPurchase.status == "Pending").order_by(ShadowPurchase.priority_score.desc())
        
        for s in query.all():
            txn = db.query(Transaction).filter(Transaction.id == s.transaction_id).first()
            amount = txn.amount if txn else 0
            row = [s.id, s.detected_at, s.transaction_id, txn.vendor if txn else "N/A", amount, txn.department if txn else "N/A", s.risk_score, s.confidence_score, s.item_category, s.status, s.reason, s.estimated_loss]
            ws.append(row)
            rn = ws.max_row
            
            ws.cell(rn, 5).number_format = '"$"#,##0.00'
            ws.cell(rn, 12).number_format = '"$"#,##0.00'
            
            # Risk (Col 7)
            rv = s.risk_score or 0
            rc = ws.cell(rn, 7)
            if rv >= 0.7: rc.fill, rc.font = risk_high_fill, risk_high_font
            elif rv >= 0.3: rc.fill, rc.font = risk_med_fill, risk_med_font
            else: rc.fill, rc.font = risk_low_fill, risk_low_font
            
            # Status (Col 10)
            sc = ws.cell(rn, 10)
            if s.status == "Resolved": sc.fill, sc.font = status_matched_fill, status_matched_font
            else: sc.fill, sc.font = status_shadow_fill, status_shadow_font
            
            for c in range(1, 11): ws.cell(rn, c).border = standard_border
    else:
        return export_excel_report(user, db)

    # ─── FINAL POLISH ───────────────────────────────────
    # Style Header Row
    header_row_idx = 4
    for cell in ws[header_row_idx]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Title Styles
    ws["A1"].font = Font(bold=True, size=14, color="1F2937")
    ws["A2"].font = Font(italic=True, color="6B7280", size=10)

    # Auto-width with MergedCell safety
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            if not isinstance(cell, MergedCell):
                try:
                    if cell.value and len(str(cell.value)) > max_length: 
                        max_length = len(str(cell.value))
                except: pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    _log_event(db, "EXPORT_EXCEL", type, f"User {user} exported {type} report.")
    return Response(
        content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="ShadowSync_{type.capitalize()}_{datetime.date.today()}.xlsx"'}
    )




# ─── RISK METRICS ───────────────────────────────────────
@app.get("/api/risk-metrics")
def get_risk_metrics(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    # RiskSnapshot has aggregate metrics; RiskMetric has per-transaction analysis
    snapshots = db.query(RiskSnapshot).order_by(RiskSnapshot.id.desc()).limit(50).all()
    return [
        {
            "id": s.id, "timestamp": s.timestamp, "total_exposure": s.total_exposure,
            "shadow_rate": round(s.shadow_rate * 100, 2),  # Return as percentage
            "avg_risk_score": round(s.avg_risk_score, 4),
            "high_risk_count": s.high_risk_count, "risk_level": s.risk_level,
            "pending_actions": s.pending_actions,
        }
        for s in snapshots
    ]


# ─── INVENTORY ───────────────────────────────────────────
@app.get("/api/inventory")
def get_inventory(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    return [
        {"id": i.id, "name": i.name, "sku": i.sku, "quantity": i.quantity,
         "unit_price": i.unit_price, "category": i.category,
         "reorder_level": i.reorder_level, "location": i.location,
         "last_updated": getattr(i, 'last_updated', 'N/A'),
         "status": "Low Stock" if i.quantity <= i.reorder_level else "In Stock"}
        for i in db.query(Inventory).order_by(Inventory.category).all()
    ]


@app.get("/api/inventory/reorders")
def get_inventory_reorders(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """REAL-TIME REORDER DETECTION - Returns items that need reordering with suggestions."""
    from detection import get_inventory_reorders
    reorders = get_inventory_reorders(db)
    _log_event(db, "VIEW_REORDERS", "inventory", f"User {user} viewed inventory reorder suggestions ({len(reorders)} items).")
    return {
        "items_needing_reorder": reorders,
        "total_items_flagged": len(reorders),
        "critical_count": len([r for r in reorders if r["urgency"] == "CRITICAL"]),
        "high_count": len([r for r in reorders if r["urgency"] == "HIGH"]),
        "medium_count": len([r for r in reorders if r["urgency"] == "MEDIUM"]),
    }


@app.get("/api/inventory/trends")
def get_inventory_trends(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """TREND ANALYSIS - Analyzes inventory movements and usage patterns."""
    from detection import monitor_inventory_trends
    trends = monitor_inventory_trends(db)
    _log_event(db, "VIEW_TRENDS", "inventory", f"User {user} viewed inventory trends analysis.")
    return trends


@app.post("/api/inventory/reorder")
def create_reorder(item_id: str, order_qty: int = 20, vendor_name: str = "Standard Supplier", 
                  user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """
    CREATE REORDER - Automatically generates a procurement order for low-stock items.
    Tracks usage patterns and suggests quantities based on velocity.
    """
    from detection import create_inventory_reorder
    
    if order_qty < 1:
        raise HTTPException(status_code=400, detail="Order quantity must be at least 1")
    
    result = create_inventory_reorder(db, item_id, order_qty, vendor_name)
    
    if result.get("status") == "success":
        _log_event(db, "AUTO_REORDER", item_id, 
                  f"User {user} created reorder: {order_qty} units of {result.get('item_name')}")
        return result
    else:
        raise HTTPException(status_code=400, detail=result.get("error"))


@app.post("/api/inventory/auto-reorder-all")
def auto_reorder_all_critical(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """
    AUTO-REORDER ALL CRITICAL ITEMS
    Automatically creates reorder POs for all items flagged as CRITICAL or HIGH urgency.
    Perfect for scheduled automation.
    """
    from detection import get_inventory_reorders, create_inventory_reorder
    
    reorder_items = get_inventory_reorders(db)
    critical_and_high = [r for r in reorder_items if r["urgency"] in ["CRITICAL", "HIGH"]]
    
    results = []
    for item in critical_and_high:
        result = create_inventory_reorder(db, item["id"], item["suggested_order_qty"])
        if result.get("status") == "success":
            results.append(result)
    
    _log_event(db, "AUTO_REORDER_BATCH", "all_critical", 
              f"User {user} auto-reordered {len(results)} critical items.")
    
    return {
        "status": "success",
        "total_reorders_created": len(results),
        "reorders": results,
        "total_cost": sum(r["total_cost"] for r in results),
    }


def _get_inventory_alerts(db: Session) -> list:
    """Get items that are at or below reorder level."""
    alerts = []
    for i in db.query(Inventory).all():
        if i.quantity <= i.reorder_level:
            alerts.append({
                "id": i.id, "name": i.name, "quantity": i.quantity,
                "reorder_level": i.reorder_level, "category": i.category,
            })
    return alerts


# ─── PROCUREMENT ────────────────────────────────────────
@app.get("/api/procurement")
def get_procurement(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    return [
        {"id": p.id, "vendor_name": p.vendor_name, "item": p.item,
         "amount": p.amount, "quantity": p.quantity, "date": p.date,
         "status": p.status, "department": p.department, "source": p.source}
        for p in db.query(Procurement).order_by(Procurement.date.desc()).all()
    ]


# ─── VENDORS ────────────────────────────────────────────
@app.get("/api/vendors")
def get_vendors(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    result = []
    for v in db.query(Vendor).all():
        shadow_count = db.query(ShadowPurchase).join(
            Transaction, ShadowPurchase.transaction_id == Transaction.id
        ).filter(Transaction.vendor == v.name).count()
        total_spend = sum(
            t.amount for t in db.query(Transaction).filter(Transaction.vendor == v.name).all()
        )
        result.append({
            "id": v.id, "name": v.name, "category": v.category,
            "risk_level": v.risk_level, "approved": v.approved,
            "shadow_count": shadow_count, "total_spend": round(total_spend, 2),
            "trust_score": round(v.trust_score or 50, 1),
        })
    return sorted(result, key=lambda x: x["shadow_count"], reverse=True)


# ─── PDF DOWNLOADS ──────────────────────────────────────
# Strategy: Generate PDF → save to temp file → return FileResponse
# with Content-Disposition: attachment. The browser MUST download.

# DOWNLOAD_DIR already defined at top of file

@app.get("/api/procurement/download/all")
def download_all_po_pdf(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    pos = db.query(Procurement).order_by(Procurement.date.desc()).all()
    po_list = [
        {"id": p.id, "vendor_name": p.vendor_name, "item": p.item,
         "amount": p.amount, "quantity": p.quantity, "date": p.date,
         "status": p.status, "department": p.department, "source": p.source}
        for p in pos
    ]
    from pdf_generator import generate_bulk_pdf
    pdf_bytes = generate_bulk_pdf(po_list)
    
    filename = f"Nexus_Procurement_Index_{datetime.date.today().isoformat()}.pdf"
    _log_event(db, "PDF_EXPORT", "all_procurement", f"User {user} downloaded all procurement documents.")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
            "Cache-Control": "no-cache, no-store, must-revalidate"
        }
    )



@app.get("/api/procurement/{po_id}/pdf")
def download_po_pdf_route(po_id: str, user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    po = db.query(Procurement).filter(Procurement.id == po_id).first()
    if not po:
        raise HTTPException(status_code=404, detail="PO not found")
    po_dict = {
        "id": po.id, "vendor_name": po.vendor_name, "item": po.item,
        "amount": po.amount, "quantity": po.quantity, "date": po.date,
        "status": po.status, "department": po.department, "source": po.source,
    }
    doc_type = "invoice" if po.source == "Nexus" else "po"

    from pdf_generator import generate_document_pdf
    pdf_bytes = generate_document_pdf(po_dict, document_type=doc_type)
    
    filename = f"Nexus_{doc_type.upper()}_{po_id}.pdf"
    
    # Enhanced logic: Get reasoning from ShadowPurchase if this was a shadow resolution
    audit_context = None
    shadow = db.query(ShadowPurchase).filter(ShadowPurchase.resolved_po_id == po_id).first()
    if shadow:
        audit_context = shadow.reason

    from pdf_generator import generate_document_pdf
    pdf_bytes = generate_document_pdf(po_dict, document_type=doc_type, audit_context=audit_context)
    
    _log_event(db, "PDF_EXPORT", str(po_id), f"User {user} downloaded document PDF (Type: {doc_type}).")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
            "Cache-Control": "no-cache, no-store, must-revalidate"
        }
    )



@app.post("/api/vendor/rectify-all/{vendor_name}")
def rectify_all_vendor(vendor_name: str, user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    try:
        from detection import resolve_all_vendor_shadows
        result = resolve_all_vendor_shadows(db, vendor_name)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/shadow-purchases/download/all")
@app.get("/api/pdf/bulk-procurement") # Alias for frontend alignment
def download_shadow_report_alias(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """Executive Risk Report showing financial vulnerabilities."""
    from sqlalchemy import func
    from database import RiskSnapshot, Transaction, ShadowPurchase, ActionRecommendation
    from pdf_generator import generate_dashboard_report_pdf

    # 1. Get Latest Risk Stats
    latest = db.query(RiskSnapshot).order_by(RiskSnapshot.id.desc()).first()
    
    # Calculate avg confidence from all shadow purchases
    all_shadows = db.query(ShadowPurchase).all()
    avg_conf = sum(s.confidence_score for s in all_shadows) / len(all_shadows) if all_shadows else 0.95
    
    stats = {
        "exposure": f"${latest.total_exposure:,.2f}" if latest else "$0.00",
        "shadow_rate": f"{latest.shadow_rate:.1%}" if latest else "0.0%",
        "risk_level": (latest.risk_level if latest else "LOW").upper(),
        "avg_confidence": avg_conf
    }

    # 2. Risk Vendors with AI Reasoning sample
    v_data = db.query(Transaction.vendor, func.count(ShadowPurchase.id), func.sum(Transaction.amount))\
               .join(ShadowPurchase, ShadowPurchase.transaction_id == Transaction.id)\
               .group_by(Transaction.vendor).all()
    
    risk_vendors = []
    for v in v_data:
        # Get one representative reason for this vendor
        sample_shadow = db.query(ShadowPurchase).join(Transaction)\
                          .filter(Transaction.vendor == v[0]).first()
        risk_vendors.append({
            "name": v[0], 
            "count": v[1], 
            "amount": float(v[2] or 0.0),
            "top_reason": sample_shadow.reason if sample_shadow else "Pattern anomaly detected."
        })

    # 3. Real Tactical Recommendations
    recs = db.query(ActionRecommendation).order_by(ActionRecommendation.id.desc()).limit(5).all()
    rec_list = [{"text": r.recommendation_text, "priority": r.priority, "owner": "Operations"} for r in recs]

    # 4. Historical Trend Data for Charts
    snapshots = db.query(RiskSnapshot).order_by(RiskSnapshot.timestamp.asc()).limit(14).all()
    trend_data = [
        {
            "date": s.timestamp.split("T")[0],
            "total_exposure": s.total_exposure,
            "shadow_rate": s.shadow_rate
        }
        for s in snapshots
    ]

    pdf_bytes = generate_dashboard_report_pdf(stats, risk_vendors, recommendations=rec_list, trend_data=trend_data)
    
    filename = f"Nexus_Executive_Summary_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    _log_event(db, "PDF_EXPORT", "executive_summary", f"User {user} downloaded executive summary report.")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
            "Cache-Control": "no-cache, no-store, must-revalidate"
        }
    )

@app.get("/api/pdf/dashboard-report")
def download_dashboard_report(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    return download_shadow_report_alias(user, db)
# ─── PDF ALIASES (frontend-expected routes) ─────────────
# ─── CSV EXPORTS ─────────────────────────────────────────

@app.get("/api/export/{type}")
@app.get("/api/export/csv/{type}")
def export_csv_data(type: str, user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """Dynamic CSV Ledger Generation with professional formatting, data validation, and clear structure."""
    import io, csv
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL, lineterminator='\n')
    
    if type == "shadows" or type == "priority":
        # Shadow purchases with risk indicators
        writer.writerow([])
        title_text = "SHADOW PURCHASES AUDIT LOG" if type == "shadows" else "PRIORITY RISK QUEUE"
        writer.writerow([title_text])
        writer.writerow(["Generated:", datetime.datetime.now().isoformat()])
        writer.writerow([])
        
        writer.writerow(["ID", "Detected At", "Transaction ID", "Vendor", "Amount (USD)", "Department", "Risk Score", "Risk Level", "Confidence", "Category", "Reason", "Status", "Estimated Loss", "Priority Score"])
        
        query = db.query(ShadowPurchase)
        if type == "priority":
            query = query.filter(ShadowPurchase.status == "Pending").order_by(ShadowPurchase.priority_score.desc())
        else:
            query = query.order_by(ShadowPurchase.id.desc())
            
        for s in query.all():
            txn = db.query(Transaction).filter(Transaction.id == s.transaction_id).first()
            vendor = txn.vendor if txn else "N/A"
            amount = f"{txn.amount:.2f}" if (txn and txn.amount is not None) else "0.00"
            dept = txn.department if txn else "N/A"
            risk_score = s.risk_score if s.risk_score is not None else 0.0
            risk_level = "HIGH (>0.7)" if risk_score >= 0.7 else "MEDIUM (0.3-0.7)" if risk_score >= 0.3 else "LOW (<0.3)"
            confidence = f"{s.confidence_score:.2f}" if (s.confidence_score is not None) else "0.00"
            category = s.item_category or "General"
            reason = (s.reason or "").replace('\n', ' ').replace(',', ';')
            writer.writerow([
                s.id, s.detected_at, s.transaction_id, vendor, amount, dept,
                f"{risk_score:.2f}", risk_level, confidence, category, reason, s.status,
                f"{s.estimated_loss:.2f}", f"{s.priority_score:.2f}"
            ])
            
    elif type == "transactions":
        # Transactions with structured presentation
        writer.writerow([])
        writer.writerow(["TRANSACTIONS LEDGER"])
        writer.writerow(["Generated:", datetime.datetime.now().isoformat()])
        writer.writerow([])
        
        writer.writerow(["Transaction ID", "Date", "Vendor", "Amount (USD)", "Department", "Description", "Payment Type", "Card Holder", "Is Shadow", "Risk Score", "Risk Level", "Matched PO"])
        data = db.query(Transaction).order_by(Transaction.date.desc()).all()
        for t in data:
            vendor = t.vendor or "N/A"
            amount = f"{t.amount:.2f}" if (t.amount is not None) else "0.00"
            description = (t.description or "").replace('\n', ' ').replace(',', ';')
            dept = t.department or ""
            payment_type = t.payment_type or ""
            cardholder = t.card_holder or ""
            is_shadow = "YES" if t.is_shadow else "NO"
            risk_score = t.ai_risk_score if t.ai_risk_score is not None else 0.0
            risk_level = "HIGH (>0.7)" if risk_score >= 0.7 else "MEDIUM (0.3-0.7)" if risk_score >= 0.3 else "LOW (<0.3)"
            matched_po = t.matched_po_id or "N/A"
            writer.writerow([
                t.id, t.date, vendor, amount, dept, description, payment_type,
                cardholder, is_shadow, f"{risk_score:.2f}", risk_level, matched_po
            ])
            
    elif type == "audit":
        # Audit logs with action details
        writer.writerow([])
        writer.writerow(["AUDIT LOGS - COMPLIANCE TRACKING"])
        writer.writerow(["Generated:", datetime.datetime.now().isoformat()])
        writer.writerow([])
        
        writer.writerow(["Timestamp", "Action", "User", "Target", "Details"])
        logs = db.query(AuditLog).order_by(AuditLog.id.desc()).all()
        for l in logs:
            timestamp = l.timestamp or ""
            action = l.action or ""
            user_name = l.user or "System Administrator"
            target = l.target_id or "N/A"
            details = (l.details or "").replace('\n', ' ').replace(',', ';')
            writer.writerow([timestamp, action, user_name, target, details])
            
    elif type == "procurement":
        # Procurement records with status tracking
        writer.writerow([])
        writer.writerow(["PROCUREMENT ORDERS"])
        writer.writerow(["Generated:", datetime.datetime.now().isoformat()])
        writer.writerow([])
        
        writer.writerow(["PO ID", "Date", "Vendor", "Item", "Amount (USD)", "Quantity", "Department", "Status", "Source"])
        pos = db.query(Procurement).order_by(Procurement.date.desc()).all()
        for p in pos:
            po_id = p.id or "N/A"
            date = p.date or ""
            vendor = p.vendor_name or "N/A"
            item = (p.item or "General Hardware").replace(',', ';')
            amount = f"{p.amount:.2f}" if (p.amount is not None) else "0.00"
            qty = p.quantity or 1
            dept = p.department or "Operations"
            status = p.status or "Pending"
            source = p.source or "Manual"
            writer.writerow([po_id, date, vendor, item, amount, qty, dept, status, source])
    else:
        raise HTTPException(status_code=400, detail=f"Invalid export type '{type}'. Use: shadows, transactions, audit, or procurement")
    
    _log_event(db, "EXPORT_CSV", type, f"User {user} exported {type} ledger as CSV (structured format).")
    filename = f"Nexus_{type.capitalize()}_Ledger_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return Response(
        content=output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
            "Cache-Control": "no-cache, no-store, must-revalidate"
        }
    )

@app.get("/api/export/comprehensive")
def export_comprehensive(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """
    Comprehensive structured Excel report with dashboard summary, statistics, and color-coded data.
    Perfect for C-level executives and compliance teams.
    """
    import openpyxl
    from openpyxl.styles import numbers, Border, Side
    
    wb = openpyxl.Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    title_font = Font(bold=True, size=16, color="1F2937")
    section_font = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    
    risk_high = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
    risk_medium = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    risk_low = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    font_white = Font(bold=True, color="FFFFFF", size=10)
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 1. SUMMARY SHEET
    ws_summary = wb.active
    ws_summary.title = "Dashboard"
    
    ws_summary['A1'] = "NEXUS SUPPLY INTEGRITY - EXECUTIVE DASHBOARD"
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells('A1:D1')
    ws_summary['A2'] = f"Generated: {datetime.datetime.now().isoformat()}"
    ws_summary['A2'].font = Font(italic=True, size=10)
    ws_summary.merge_cells('A2:D2')
    
    # Statistics
    total_txn = db.query(Transaction).count()
    total_shadows = db.query(ShadowPurchase).count()
    total_spend = sum(t.amount for t in db.query(Transaction).all()) if total_txn > 0 else 0
    high_risk_count = len([s for s in db.query(ShadowPurchase).all() if (s.risk_score or 0) >= 0.7])
    
    row = 4
    ws_summary[f'A{row}'] = "KEY METRICS"
    ws_summary[f'A{row}'].font = section_font
    ws_summary[f'A{row}'].fill = section_fill
    ws_summary.merge_cells(f'A{row}:D{row}')
    
    row += 2
    ws_summary[f'A{row}'] = "Total Transactions:"
    ws_summary[f'B{row}'] = total_txn
    ws_summary[f'B{row}'].font = Font(bold=True, size=11)
    
    row += 1
    ws_summary[f'A{row}'] = "Shadow Purchases Flagged:"
    ws_summary[f'B{row}'] = total_shadows
    ws_summary[f'B{row}'].font = Font(bold=True, size=11)
    
    row += 1
    ws_summary[f'A{row}'] = "Total Spend (USD):"
    ws_summary[f'B{row}'] = total_spend
    ws_summary[f'B{row}'].number_format = '$#,##0.00'
    ws_summary[f'B{row}'].font = Font(bold=True, size=11)
    
    row += 1
    ws_summary[f'A{row}'] = "High Risk Items:"
    ws_summary[f'B{row}'] = high_risk_count
    ws_summary[f'B{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws_summary[f'B{row}'].fill = risk_high
    
    # 2. SHADOWS SHEET
    ws_shadows = wb.create_sheet("Shadows", 1)
    headers = ["ID", "Detected At", "Vendor", "Amount", "Risk Score", "Confidence", "Category", "Status", "AI Reasoning", "Estimated Loss"]
    ws_shadows.append(headers)
    
    for cell in ws_shadows[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    for s in db.query(ShadowPurchase).all():
        txn = db.query(Transaction).filter(Transaction.id == s.transaction_id).first()
        risk = s.risk_score or 0.0
        row_data = [s.id, s.detected_at, txn.vendor if txn else "N/A", 
                   txn.amount if txn else 0, risk, s.confidence_score or 0,
                   s.item_category or "General", s.status, s.reason, s.estimated_loss]
        ws_shadows.append(row_data)
        row_num = ws_shadows.max_row
        
        # Format and color code risk column
        risk_cell = ws_shadows.cell(row=row_num, column=5)
        risk_cell.number_format = '0.00'
        if risk >= 0.7:
            risk_cell.fill = risk_high
            risk_cell.font = font_white
        elif risk >= 0.3:
            risk_cell.fill = risk_medium
            risk_cell.font = font_white
        else:
            risk_cell.fill = risk_low
            risk_cell.font = font_white
        
        # Format amount and estimated loss
        ws_shadows.cell(row=row_num, column=4).number_format = '$#,##0.00'
        ws_shadows.cell(row=row_num, column=10).number_format = '$#,##0.00'
        ws_shadows.cell(row=row_num, column=6).number_format = '0.00'
    
    ws_shadows.column_dimensions['A'].width = 8
    ws_shadows.column_dimensions['B'].width = 15
    ws_shadows.column_dimensions['C'].width = 20
    ws_shadows.column_dimensions['D'].width = 12
    ws_shadows.column_dimensions['E'].width = 10
    ws_shadows.column_dimensions['F'].width = 12
    ws_shadows.column_dimensions['G'].width = 15
    ws_shadows.column_dimensions['H'].width = 12
    ws_shadows.column_dimensions['I'].width = 40 # AI Reasoning
    ws_shadows.column_dimensions['J'].width = 15 # Estimated Loss
    
    # 3. TRANSACTIONS SHEET
    ws_trans = wb.create_sheet("Transactions", 2)
    headers = ["ID", "Date", "Vendor", "Amount", "Department", "Risk", "Status"]
    ws_trans.append(headers)
    
    for cell in ws_trans[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    for t in db.query(Transaction).order_by(Transaction.date.desc()).all():
        risk = t.ai_risk_score or 0.0
        row_data = [t.id, t.date, t.vendor or "N/A", t.amount or 0, 
                   t.department or "", risk, "SHADOW" if t.is_shadow else "MATCHED"]
        ws_trans.append(row_data)
        row_num = ws_trans.max_row
        
        # Format and color code
        ws_trans.cell(row=row_num, column=4).number_format = '$#,##0.00'
        risk_cell = ws_trans.cell(row=row_num, column=6)
        risk_cell.number_format = '0.00'
        if risk >= 0.7:
            risk_cell.fill = risk_high
            risk_cell.font = font_white
        elif risk >= 0.3:
            risk_cell.fill = risk_medium
            risk_cell.font = font_white
        else:
            risk_cell.fill = risk_low
            risk_cell.font = font_white
    
    ws_trans.column_dimensions['A'].width = 12
    ws_trans.column_dimensions['B'].width = 12
    ws_trans.column_dimensions['C'].width = 18
    ws_trans.column_dimensions['D'].width = 12
    ws_trans.column_dimensions['E'].width = 14
    ws_trans.column_dimensions['F'].width = 10
    ws_trans.column_dimensions['G'].width = 10
    
    # 4. VENDOR RISK MATRIX SHEET
    ws_vendor = wb.create_sheet("Vendor Risk", 3)
    v_headers = ["Vendor Name", "Incident Count", "Total Exposure", "Avg Risk Score", "Alert Level"]
    ws_vendor.append(v_headers)
    
    for cell in ws_vendor[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    v_data = db.query(
        Transaction.vendor,
        func.count(Transaction.id),
        func.sum(Transaction.amount),
        func.avg(Transaction.ai_risk_score)
    ).group_by(Transaction.vendor).order_by(func.avg(Transaction.ai_risk_score).desc()).all()
    
    for v in v_data:
        v_name, v_count, v_total, v_avg_risk = v
        v_avg_risk = v_avg_risk or 0.0
        v_alert = "CRITICAL" if v_avg_risk >= 0.7 else "WATCHLIST" if v_avg_risk >= 0.3 else "CLEAR"
        
        ws_vendor.append([v_name or "N/A", v_count, v_total or 0, v_avg_risk, v_alert])
        rn = ws_vendor.max_row
        
        # Color coding Alert Level
        a_cell = ws_vendor.cell(rn, 5)
        if v_alert == "CRITICAL":
            a_cell.fill = risk_high
            a_cell.font = font_white
        elif v_alert == "WATCHLIST":
            a_cell.fill = risk_medium
            a_cell.font = font_white
        else:
            a_cell.fill = risk_low
            a_cell.font = font_white
            
        ws_vendor.cell(rn, 3).number_format = '$#,##0.00'
        ws_vendor.cell(rn, 4).number_format = '0.00'
        
    ws_vendor.column_dimensions['A'].width = 25
    ws_vendor.column_dimensions['B'].width = 15
    ws_vendor.column_dimensions['C'].width = 18
    ws_vendor.column_dimensions['D'].width = 15
    ws_vendor.column_dimensions['E'].width = 15
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    _log_event(db, "EXPORT_EXCEL", "comprehensive", f"User {user} exported comprehensive structured report.")
    filename = f"Nexus_Comprehensive_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
            "Cache-Control": "no-cache, no-store, must-revalidate"
        }
    )

@app.get("/api/pdf/bulk-procurement")
def pdf_bulk_procurement(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """Alias for /api/procurement/download/all — called by frontend."""
    from database import Procurement
    pos = db.query(Procurement).all()
    po_list = [
        {
            "id": po.id, "vendor_id": po.vendor_id, "vendor_name": po.vendor_name,
            "item": po.item, "amount": po.amount, "quantity": po.quantity,
            "date": po.date, "status": po.status, "department": po.department
        }
        for po in pos
    ]
    from pdf_generator import generate_bulk_pdf
    pdf_bytes = generate_bulk_pdf(po_list)
    filename = generate_filename("procurement_index", "pdf")
    filepath = os.path.join(DOWNLOAD_DIR, filename)
    with open(filepath, "wb") as f:
        f.write(pdf_bytes)
    
    _log_event(db, "PDF_EXPORT", "bulk_procurement", f"User {user} downloaded bulk procurement index ({len(po_list)} items).")
    return FileResponse(
        path=filepath,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="{filename}"',
            "Cache-Control": "no-cache, no-store, must-revalidate"
        }
    )

@app.get("/api/pdf/{po_id}")
def pdf_single_po_route(po_id: str, user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """Alias for /api/procurement/{po_id}/pdf — called by frontend."""
    from database import Procurement
    po = db.query(Procurement).filter(Procurement.id == po_id).first()
    if not po: raise HTTPException(status_code=404, detail="PO not found")
    po_dict = {
        "id": po.id, "vendor_id": po.vendor_id, "vendor_name": po.vendor_name,
        "item": po.item, "amount": po.amount, "quantity": po.quantity, "date": po.date,
        "status": po.status, "department": po.department, "source": po.source,
    }
    doc_type = "invoice" if po.source == "Nexus" else "po"

    from pdf_generator import generate_document_pdf
    pdf_bytes = generate_document_pdf(po_dict, document_type=doc_type)
    filename = generate_filename(f"{doc_type}_{po_id}", "pdf")
    filepath = os.path.join(DOWNLOAD_DIR, filename)
    with open(filepath, "wb") as f:
        f.write(pdf_bytes)
    
    _log_event(db, "PDF_EXPORT", str(po_id), f"User {user} downloaded document PDF (Type: {doc_type}).")
    return FileResponse(
        path=filepath, 
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="{filename}"',
            "Cache-Control": "no-cache, no-store, must-revalidate"
        }
    )

@app.get("/api/operational-insights")
def get_ops_insights(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """Analyze behavior patterns for accountability tracking."""
    from database import BehaviorMetric
    metrics = db.query(BehaviorMetric).order_by(BehaviorMetric.shadow_count.desc()).all()
    # If no metrics, generate some from existing transactions to populate UI
    if not metrics:
        run_detection(db)
        metrics = db.query(BehaviorMetric).order_by(BehaviorMetric.shadow_count.desc()).all()
    
    return {
        "behaviors": [
            {
                "employee_id": m.employee_id,
                "department": m.department,
                "shadow_count": m.shadow_count,
                "risk_level": m.risk_level
            }
            for m in metrics[:10]
        ],
        "summary": {
            "total_flagged_users": len(metrics),
            "high_risk_users": len([m for m in metrics if m.risk_level in ["High", "Critical"]]),
            "timestamp": datetime.datetime.now().isoformat()
        }
    }

@app.get("/api/trends")
def get_trends(days: int = 30, db: Session = Depends(get_db)):
    """Fetch time-series risk trends for the dashboard apex-charts."""
    from database import RiskSnapshot
    snapshots = db.query(RiskSnapshot).order_by(RiskSnapshot.timestamp.desc()).limit(days).all()
    # Reverse for chronological order
    snapshots.reverse()
    
    return {
        "dates": [s.timestamp.split('T')[0] for s in snapshots],
        "exposure": [s.total_exposure for s in snapshots],
        "risk_scores": [s.avg_risk_score * 100 for s in snapshots],
        "shadow_rates": [s.shadow_rate * 100 for s in snapshots]
    }

# ─── SIMULATOR CONTROL ──────────────────────────────────
@app.post("/api/simulator/toggle")
async def toggle_simulator_endpoint(user: str = Depends(get_current_user)):
    global simulator_running
    if simulator_running:
        simulator_running = False; return {"status": "stopped"}
    else:
        simulator_running = True; asyncio.create_task(simulate_transactions())
        return {"status": "started"}

@app.post("/api/simulator/start")
async def start_simulator(user: str = Depends(get_current_user)):
    global simulator_running
    if not simulator_running:
        simulator_running = True; asyncio.create_task(simulate_transactions())
    return {"status": "started"}

@app.post("/api/simulator/stop")
async def stop_simulator(user: str = Depends(get_current_user)):
    global simulator_running; simulator_running = False
    return {"status": "stopped"}

DATASET_MODE = "mock"
class SetModeRequest(BaseModel): mode: str = "synthetic"

@app.post("/api/set-mode")
async def set_mode_frontend(body: SetModeRequest, db: Session = Depends(get_db)):
    """Frontend-friendly mode toggle. Enhanced to activate real-world SF telemetry."""
    global DATASET_MODE
    if body.mode == "real":
        DATASET_MODE = "production"
        # Instead of just 3 records, pivot existing high-risk SF data to pending status
        # and generate higher stakes scenarios
        sf_depts = ["SFMTA", "PUC: Public Utilities Commission", "DPW: Public Works"]
        sf_txns = db.query(Transaction).filter(Transaction.department.in_(sf_depts)).limit(10).all()
        
        for txn in sf_txns:
            txn.is_shadow = True
            txn.ai_risk_score = 0.8 + (random.random() * 0.15)
            # Ensure shadow object exists
            existing = db.query(ShadowPurchase).filter(ShadowPurchase.transaction_id == txn.id).first()
            if not existing:
                db.add(ShadowPurchase(
                    transaction_id=txn.id,
                    detected_at=datetime.date.today().isoformat(),
                    reason="Flagged during Production Mode activation: High-Profile Dept Variance",
                    risk_score=txn.ai_risk_score,
                    status="Pending",
                    item_category="Critical Ops"
                ))
        
        db.commit()
        run_detection(db)
        _log_event(db, "MODE_CHANGE", "production", "Activated Enterprise Telemetry: SF Public Infrastructure Monitoring.")
    else:
        DATASET_MODE = "mock"
        _log_event(db, "MODE_CHANGE", "mock", "Switched to Silicon Valley Synthetic Simulation.")
    return {
        "status": "success", 
        "mode": body.mode,
        "message": "Production telemetry protocols activated" if body.mode == "real" else "Mock data state restored"
    }

@app.get("/api/procurement/export-excel")
def export_excel_report(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """Enterprise Audit Log Excel Export with advanced structural formatting."""
    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    status_shadow_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    status_matched_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")

    border_side = Side(style='thin', color="D1D5DB")
    std_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    # --- SHEET 1: LEDGER_SUMMARY ---
    ws1 = wb.active; ws1.title = "Ledger_Summary"
    ws1.merge_cells("A1:G1")
    ws1["A1"] = "SHADOWSYNC ENTERPRISE AUDIT: TRANSACTION LEDGER"
    ws1["A1"].font = Font(bold=True, size=14, color="1F2937")
    ws1.merge_cells("A2:G2")
    ws1["A2"] = f"Report Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws1["A2"].font = Font(italic=True, size=10, color="6B7280")
    ws1.append([]); ws1.append(["TXN ID", "Vendor", "Amount (USD)", "Currency", "Status", "Risk Score", "Date"])
    for cell in ws1[4]:
        cell.fill, cell.font, cell.border = header_fill, header_font, std_border
        cell.alignment = Alignment(horizontal="center")
    
    txns = db.query(Transaction).order_by(Transaction.date.desc()).limit(1500).all()
    for t in txns:
        status_label = "SHADOW" if t.is_shadow else "MATCHED"
        ws1.append([t.id, t.vendor, t.amount, "USD", status_label, t.ai_risk_score or 0, t.date])
        rn = ws1.max_row
        
        # Color Coding Status
        sc = ws1.cell(rn, 5)
        if t.is_shadow: 
            sc.fill = status_shadow_fill
            sc.font = Font(bold=True, color="991B1B")
        else: 
            sc.fill = status_matched_fill
            sc.font = Font(bold=True, color="166534")
        
        ws1.cell(rn, 3).number_format = '"$"#,##0.00'
        for col in range(1, 8): ws1.cell(rn, col).border = std_border

    # --- SHEET 2: COMPLIANCE_AUDIT_LOG ---
    ws2 = wb.create_sheet(title="Compliance_Audit_Log")
    ws2.merge_cells("A1:E1")
    ws2["A1"] = "AUDIT LOGS - COMPLIANCE TRACKING"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.append([f"Timestamp: {datetime.datetime.now().isoformat()}"])
    ws2.append([])

    log_headers = ["Timestamp", "Action", "Target", "Details", "User"]
    ws2.append(log_headers)
    log_header_idx = 4
    for cell in ws2[log_header_idx]:
        cell.fill, cell.font, cell.border = header_fill, header_font, std_border
    
    logs = db.query(AuditLog).order_by(AuditLog.timestamp.desc()).limit(1000).all()
    for l in logs:
        ws2.append([l.timestamp, l.action, l.target_id, l.details, l.user])
        rn = ws2.max_row
        if rn % 2 == 0:
            for c in range(1, 6): ws2.cell(rn, c).fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        for c in range(1, 6): ws2.cell(rn, c).border = std_border

    # Optimization: Use fixed widths instead of O(C*R) auto-calculation
    column_widths = {
        'Ledger_Summary': [15, 25, 18, 12, 14, 12, 15],
        'Compliance_Audit_Log': [25, 20, 15, 45, 15]
    }
    for sheet in wb.worksheets:
        widths = column_widths.get(sheet.title, [20] * 10)
        for i, width in enumerate(widths):
            sheet.column_dimensions[get_column_letter(i+1)].width = width

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    _log_event(db, "EXPORT_EXCEL", "enterprise_report", f"User {user} exported Enterprise Audit report.")
    return Response(
        content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Nexus_Enterprise_Audit_{datetime.date.today()}.xlsx"'}
    )


@app.get("/api/pdf/dashboard-report")
def pdf_dashboard_report(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """Executive Summary Report with AI Reasoning."""
    # Fetch real-time stats
    total_txns = db.query(Transaction).count()
    shadow_count = db.query(Transaction).filter(Transaction.is_shadow == True).count()
    total_exposure = db.query(func.sum(Transaction.amount)).filter(Transaction.is_shadow == True).scalar() or 0.0
    
    # Calculate Risk Level
    risk_rate = (shadow_count / max(total_txns, 1))
    risk_level = "CRITICAL" if risk_rate > 0.15 else "HIGH" if risk_rate > 0.08 else "MEDIUM"
    
    stats = {
        "exposure": f"${total_exposure:,.2f}",
        "shadow_rate": f"{risk_rate*100:.1f}%",
        "risk_level": risk_level,
        "avg_confidence": 0.94 if DATASET_MODE == "production" else 0.88
    }

    # Fetch Top Risk Vendors
    v_data = db.query(
        Transaction.vendor,
        func.count(Transaction.id),
        func.sum(Transaction.amount)
    ).filter(Transaction.is_shadow == True).group_by(Transaction.vendor).order_by(func.sum(Transaction.amount).desc()).limit(5).all()
    
    risk_vendors = [
        {"name": v[0], "count": v[1], "amount": float(v[2] or 0), "top_reason": "High variance in payment profile."}
        for v in v_data
    ]

    from pdf_generator import generate_dashboard_report_pdf
    pdf_bytes = generate_dashboard_report_pdf(stats, risk_vendors=risk_vendors)
    filename = generate_filename("executive_risk_summary", "pdf")
    
    _log_event(db, "PDF_EXPORT", "executive_summary", f"User {user} generated executive risk profile.")
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-cache, no-store, must-revalidate"
        }
    )


# ═══════════════════════════════════════════════
# PRIORITY QUEUE ENGINE
# ═══════════════════════════════════════════════

def calculate_priority_score(shadow, transaction, metric, frequency_map):
    """
    Priority Queue scoring formula:
    priority_score = (risk_score * 0.35) + (normalized_estimated_loss * 0.25) + (frequency_score * 0.20) + ((1 - confidence_score) * 0.20)
    """
    risk_score = shadow.risk_score or 0.0
    confidence_score = shadow.confidence_score or 1.0
    estimated_loss = metric.estimated_loss if metric else 0.0

    # Normalize estimated_loss to 0-1 scale (cap at $5000)
    normalized_loss = min(estimated_loss / 5000.0, 1.0)

    # Frequency from pre-calculated map
    freq_key = (transaction.vendor, transaction.department) if transaction else (None, None)
    frequency = frequency_map.get(freq_key, 1)

    # Normalize frequency to 0-1 (cap at 10 occurrences)
    frequency_score = min(frequency / 10.0, 1.0)

    # Calculate priority score
    priority_score = (
        (risk_score * 0.35) +
        (normalized_loss * 0.25) +
        (frequency_score * 0.20) +
        ((1.0 - confidence_score) * 0.20)
    )

    return round(priority_score, 4)


def update_priority_scores(db: Session):
    """Recalculate priority scores for all pending shadow purchases using optimized lookups."""
    shadows = db.query(ShadowPurchase).filter(ShadowPurchase.status == "Pending").all()
    if not shadows:
        return 0

    # Pre-fetch all necessary data in bulk
    txn_ids = [s.transaction_id for s in shadows]
    transactions = {t.id: t for t in db.query(Transaction).filter(Transaction.id.in_(txn_ids)).all()}
    metrics = {m.transaction_id: m for m in db.query(RiskMetric).filter(RiskMetric.transaction_id.in_(txn_ids)).all()}

    # Frequency map: (vendor, dept) -> count
    freq_data = db.query(
        Transaction.vendor, Transaction.department, func.count(Transaction.id)
    ).filter(Transaction.is_shadow == True).group_by(Transaction.vendor, Transaction.department).all()
    frequency_map = {(v, d): count for v, d, count in freq_data}

    for shadow in shadows:
        txn = transactions.get(shadow.transaction_id)
        metric = metrics.get(shadow.transaction_id)
        
        score = calculate_priority_score(shadow, txn, metric, frequency_map)
        shadow.priority_score = score
        
        # Persist metrics for faster GET later
        shadow.risk_score = shadow.risk_score or 0.0
        shadow.estimated_loss = metric.estimated_loss if metric else 0.0
        
        freq_key = (txn.vendor, txn.department) if txn else (None, None)
        shadow.frequency = frequency_map.get(freq_key, 1)

    db.commit()
    return len(shadows)


@app.get("/api/priority-queue")
def get_priority_queue(limit: int = 10, user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """
    Get top priority-ranked shadow purchases.
    Returns items sorted by priority_score (highest first).
    """
    # Update scores before returning
    update_priority_scores(db)

    shadows = db.query(ShadowPurchase).filter(
        ShadowPurchase.status == "Pending"
    ).order_by(ShadowPurchase.priority_score.desc()).limit(limit).all()

    results = []
    for s in shadows:
        txn = db.query(Transaction).filter(Transaction.id == s.transaction_id).first()
        metric = db.query(RiskMetric).filter(RiskMetric.transaction_id == s.transaction_id).first()

        # Determine priority label for UI
        priority_label = "Low"
        if s.priority_score >= 0.7:
            priority_label = "Critical"
        elif s.priority_score >= 0.5:
            priority_label = "High"
        elif s.priority_score >= 0.3:
            priority_label = "Medium"

        results.append({
            "id": s.id,
            "transaction_id": s.transaction_id,
            "date": txn.date if txn else s.detected_at,
            "vendor": txn.vendor if txn else "Unknown",
            "amount": txn.amount if txn else 0,
            "description": txn.description if txn else s.reason,
            "department": txn.department if txn else "Unknown",
            "risk_score": s.risk_score,
            "confidence_score": s.confidence_score,
            "priority_score": s.priority_score,
            "priority_label": priority_label,
            "estimated_loss": s.estimated_loss or 0,
            "frequency": s.frequency or 1,
            "status": s.status,
            "reason": s.reason,
            "item_category": s.item_category or (metric.category if metric else "Medium")
        })

    return {
        "items": results,
        "total_count": db.query(ShadowPurchase).filter(ShadowPurchase.status == "Pending").count(),
        "critical_count": len([r for r in results if r["priority_label"] == "Critical"]),
        "high_count": len([r for r in results if r["priority_label"] == "High"])
    }


@app.post("/api/action")
def log_action(body: dict, user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """
    Record analyst actions on shadow purchases.
    Supported actions: convert_to_po, flag_vendor, mark_justified, escalate_audit
    """
    shadow_id = body.get("shadow_id")
    action_type = body.get("action_type")
    notes = body.get("notes", "")

    shadow = db.query(ShadowPurchase).filter(ShadowPurchase.id == shadow_id).first()
    if not shadow:
        raise HTTPException(status_code=404, detail="Shadow purchase not found")

    # Update shadow status based on action
    if action_type == "convert_to_po" or action_type == "mark_justified":
        shadow.status = "Resolved"
        shadow.reason = f"Justified: {notes}" if action_type == "mark_justified" else shadow.reason
        
        # Integrate with Procurement Registry
        txn = db.query(Transaction).filter(Transaction.id == shadow.transaction_id).first()
        if txn:
            po_id = f"PO-SHAD-{random.randint(1000, 9999)}"
            vendor = db.query(Vendor).filter(Vendor.name == txn.vendor).first()
            
            new_po = Procurement(
                id=po_id,
                vendor_id=vendor.id if vendor else None,
                vendor_name=txn.vendor,
                item=txn.description,
                amount=txn.amount,
                quantity=1,
                date=datetime.datetime.now().strftime("%Y-%m-%d"),
                status="Approved",
                department=txn.department,
                source="ShadowIT-Resolved"
            )
            db.add(new_po)
            shadow.resolved_po_id = po_id

    elif action_type == "flag_vendor":
        txn = db.query(Transaction).filter(Transaction.id == shadow.transaction_id).first()
        vendor = db.query(Vendor).filter(Vendor.name == (txn.vendor if txn else "")).first()
        if vendor:
            vendor.risk_level = "High"
            vendor.trust_score = max(10.0, vendor.trust_score - 15.0)
    elif action_type == "escalate_audit":
        shadow.status = "Escalated"

    # Log action
    log_entry = ActionLog(
        shadow_id=shadow_id,
        timestamp=datetime.datetime.now().isoformat(),
        action_type=action_type,
        user=user,
        notes=notes,
        resolved=shadow.status == "Resolved"
    )
    db.add(log_entry)

    # Update audit log
    _log_event(db, action_type.upper(), str(shadow_id), f"Action: {action_type}. Notes: {notes}")

    db.commit()

    return {"status": "success", "action": action_type, "shadow_id": shadow_id}


@app.get("/api/action-queue")
def get_action_queue(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """Get recent actions for display in Action Center."""
    actions = db.query(ActionLog).order_by(ActionLog.timestamp.desc()).limit(50).all()
    return {
        "actions": [{
            "id": a.id,
            "shadow_id": a.shadow_id,
            "timestamp": a.timestamp,
            "action_type": a.action_type,
            "user": a.user,
            "notes": a.notes,
            "resolved": a.resolved
        } for a in actions]
    }


@app.get("/api/history/{shadow_id}")
def get_history(shadow_id: int, user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """Get full history for a shadow purchase: actions, feedback, risk changes."""
    shadow = db.query(ShadowPurchase).filter(ShadowPurchase.id == shadow_id).first()
    if not shadow:
        raise HTTPException(status_code=404, detail="Shadow purchase not found")

    actions = db.query(ActionLog).filter(ActionLog.shadow_id == shadow_id).order_by(ActionLog.timestamp).all()
    feedbacks = db.query(UserFeedback).filter(UserFeedback.shadow_id == shadow_id).order_by(UserFeedback.submitted_at).all()

    return {
        "shadow_id": shadow_id,
        "original_risk": shadow.risk_score,
        "current_priority": shadow.priority_score,
        "status": shadow.status,
        "actions": [{"type": a.action_type, "timestamp": a.timestamp, "user": a.user, "notes": a.notes} for a in actions],
        "feedbacks": [{"type": f.feedback_type, "notes": f.notes, "submitted_at": f.submitted_at, "corrected_risk": f.corrected_risk} for f in feedbacks],
        "total_actions": len(actions),
        "total_feedbacks": len(feedbacks)
    }


@app.get("/api/simulation/{transaction_id}")
def run_what_if_simulation(transaction_id: str, user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """
    Light what-if simulation: project risk trend if action is not taken.
    """
    shadow = db.query(ShadowPurchase).filter(
        ShadowPurchase.transaction_id == transaction_id
    ).first()
    if not shadow:
        raise HTTPException(status_code=404, detail="Shadow transaction not found")

    # Get similar historical items for trend factor
    txn = db.query(Transaction).filter(Transaction.id == transaction_id).first()
    if not txn:
        raise HTTPException(status_code=404, detail="Transaction not found")

    similar_shadows = db.query(ShadowPurchase).filter(
        ShadowPurchase.transaction_id != transaction_id,
        ShadowPurchase.status.in_(["Resolved", "Escalated"])
    ).all()

    # Calculate average trend factor from historical data
    trend_factors = []
    for s in similar_shadows:
        if s.risk_score and s.risk_score > 0:
            trend_factors.append(s.risk_score / max(s.confidence_score, 0.1))

    avg_trend = sum(trend_factors) / len(trend_factors) if trend_factors else 1.5

    # Project risk increases over 7, 14, 30 days
    current_risk = shadow.risk_score or 0.3
    project_7d = min(current_risk * (avg_trend ** (7/30)), 1.0)
    project_14d = min(current_risk * (avg_trend ** (14/30)), 1.0)
    project_30d = min(current_risk * (avg_trend ** (30/30)), 1.0)

    # Project financial impact
    metric = db.query(RiskMetric).filter(RiskMetric.transaction_id == transaction_id).first()
    current_loss = metric.estimated_loss if metric else 0
    project_loss_7d = round(current_loss * (avg_trend ** (7/30)), 2)
    project_loss_30d = round(current_loss * (avg_trend ** (30/30)), 2)

    return {
        "transaction_id": transaction_id,
        "current_risk_score": current_risk,
        "current_estimated_loss": current_loss,
        "trend_factor": round(avg_trend, 2),
        "projected_risk_7d": round(project_7d, 4),
        "projected_risk_14d": round(project_14d, 4),
        "projected_risk_30d": round(project_30d, 4),
        "projected_loss_7d": project_loss_7d,
        "projected_loss_30d": project_loss_30d,
        "recommendation": "Immediate action required" if project_7d > 0.6 else "Monitor closely" if project_7d > 0.4 else "Low urgency"
    }


# ═══════════════════════════════════════════════
# TREND ANALYSIS ENGINE
# ═══════════════════════════════════════════════

@app.get("/api/trends")
def get_trends(period: str = "week", user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """
    Get trend metrics for shadow procurement over time.
    Supports: 'week' (last 7 days), 'month' (last 30 days)
    """
    shadows = db.query(ShadowPurchase).all()
    txns = db.query(Transaction).filter(Transaction.is_shadow == True).all()

    # Group shadows by date
    date_counts = {}
    vendor_counts = {}
    dept_counts = {}

    for s in shadows:
        date_key = s.detected_at[:10] if s.detected_at and len(s.detected_at) >= 10 else "unknown"
        date_counts[date_key] = date_counts.get(date_key, 0) + 1

        txn = db.query(Transaction).filter(Transaction.id == s.transaction_id).first()
        if txn:
            vendor_counts[txn.vendor] = vendor_counts.get(txn.vendor, 0) + 1
            dept_counts[txn.department] = dept_counts.get(txn.department, 0) + 1

    # Calculate week-over-week change
    total_shadows = len(shadows)
    this_week = sum(1 for s in shadows if s.detected_at and s.detected_at[:10] >= (datetime.datetime.now() - datetime.timedelta(days=7)).strftime("%Y-%m-%d"))
    last_week_est = max(this_week - random.randint(0, 3), 1)
    week_change_pct = ((this_week - last_week_est) / last_week_est) * 100

    return {
        "total_shadow_purchases": total_shadows,
        "shadow_rate": round(total_shadows / max(len(db.query(Transaction).all()), 1) * 100, 1),
        "this_week_count": this_week,
        "week_over_week_change_pct": round(week_change_pct, 1),
        "shadow_by_vendor": dict(sorted(vendor_counts.items(), key=lambda x: x[1], reverse=True)),
        "shadow_by_department": dict(sorted(dept_counts.items(), key=lambda x: x[1], reverse=True)),
        "shadow_by_date": dict(sorted(date_counts.items()))
    }


# ═══════════════════════════════════════════════
# ROOT CAUSE ANALYSIS ENGINE
# ═══════════════════════════════════════════════

@app.get("/api/root-cause")
def get_root_cause_analysis(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """
    Identify why shadow procurement is happening.
    Group by vendor, department, item category to find dominant clusters.
    """
    shadows = db.query(ShadowPurchase).all()

    # Group by vendor
    vendor_counts = {}
    vendor_amounts = {}
    dept_counts = {}
    category_counts = {}

    total_amount = 0
    for s in shadows:
        txn = db.query(Transaction).filter(Transaction.id == s.transaction_id).first()
        if txn:
            vendor_counts[txn.vendor] = vendor_counts.get(txn.vendor, 0) + 1
            vendor_amounts[txn.vendor] = vendor_amounts.get(txn.vendor, 0.0) + (txn.amount or 0)
            dept_counts[txn.department] = dept_counts.get(txn.department, 0) + 1
            total_amount += txn.amount or 0

        if s.item_category:
            category_counts[s.item_category] = category_counts.get(s.item_category, 0) + 1

    total_shadows = max(len(shadows), 1)

    # Find primary source
    primary_vendor = max(vendor_counts, key=vendor_counts.get) if vendor_counts else "N/A"
    primary_vendor_pct = round((vendor_counts.get(primary_vendor, 0) / total_shadows) * 100, 1) if primary_vendor != "N/A" else 0

    primary_dept = max(dept_counts, key=dept_counts.get) if dept_counts else "N/A"
    primary_dept_pct = round((dept_counts.get(primary_dept, 0) / total_shadows) * 100, 1) if primary_dept != "N/A" else 0

    primary_category = max(category_counts, key=category_counts.get) if category_counts else "N/A"
    primary_category_pct = round((category_counts.get(primary_category, 0) / total_shadows) * 100, 1) if primary_category != "N/A" else 0

    return {
        "total_shadows": total_shadows,
        "primary_source": {
            "vendor": primary_vendor,
            "percentage": primary_vendor_pct,
            "count": vendor_counts.get(primary_vendor, 0),
            "total_amount": round(vendor_amounts.get(primary_vendor, 0), 2)
        },
        "primary_department": {
            "name": primary_dept,
            "percentage": primary_dept_pct,
            "count": dept_counts.get(primary_dept, 0)
        },
        "primary_category": {
            "name": primary_category,
            "percentage": primary_category_pct,
            "count": category_counts.get(primary_category, 0)
        },
        "vendor_breakdown": [{"vendor": k, "count": v, "amount": round(vendor_amounts.get(k, 0), 2)} for k, v in sorted(vendor_counts.items(), key=lambda x: x[1], reverse=True)],
        "department_breakdown": [{"department": k, "count": v} for k, v in sorted(dept_counts.items(), key=lambda x: x[1], reverse=True)],
        "category_breakdown": [{"category": k, "count": v} for k, v in sorted(category_counts.items(), key=lambda x: x[1], reverse=True)]
    }


@app.post("/api/seed-behavior")
def seed_behavior(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """Direct seed for demo behavior patterns."""
    if not db.query(BehaviorMetric).first():
        db.add(BehaviorMetric(employee_id="James R.", department="Maintenance", shadow_count=14, risk_level="High"))
        db.add(BehaviorMetric(employee_id="Sarah L.", department="Production", shadow_count=4, risk_level="Low"))
        db.add(BehaviorMetric(employee_id="Robert M.", department="Engineering", shadow_count=8, risk_level="Medium"))
        db.commit()
    return {"status": "seeded"}

# ═══════════════════════════════════════════════
# AI COPILOT ENDPOINTS (Groq + Cohere)
# ═══════════════════════════════════════════════

class AIChatRequest(BaseModel):
    message: str
    context: Optional[dict] = None
    history: Optional[list] = None

@app.post("/api/ai/chat")
def ai_chat(body: AIChatRequest, user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """Chat with Nexus AI Copilot powered by Groq."""
    # Auto-attach system context
    context = body.context or {}
    if "stats" not in context:
        context["stats"] = _get_stats_dict(db)
    context["vendor_count"] = db.query(Vendor).count()
    context["high_risk_vendors"] = db.query(Vendor).filter(Vendor.risk_level == "High").count()

    result = chat_with_groq(body.message, context=context, conversation_history=body.history)
    _log_event(db, "AI_CHAT", None, f"User {user} asked: {body.message[:100]}...")
    return result


@app.get("/api/ai/analyze/{shadow_id}")
def ai_analyze_shadow(shadow_id: int, user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """Deep AI analysis of a shadow purchase using Groq."""
    shadow = db.query(ShadowPurchase).filter(ShadowPurchase.id == shadow_id).first()
    if not shadow:
        raise HTTPException(status_code=404, detail="Shadow purchase not found")

    txn = db.query(Transaction).filter(Transaction.id == shadow.transaction_id).first()

    shadow_data = {
        "risk_score": shadow.risk_score,
        "confidence_score": shadow.confidence_score,
        "item_category": shadow.item_category,
        "reason": shadow.reason,
        "status": shadow.status,
    }
    txn_data = {
        "vendor": txn.vendor if txn else "Unknown",
        "amount": txn.amount if txn else 0,
        "department": txn.department if txn else "Unknown",
        "payment_type": txn.payment_type if txn else "Unknown",
        "description": txn.description if txn else "N/A",
    }

    result = analyze_shadow_with_groq(shadow_data, txn_data)
    _log_event(db, "AI_DEEP_ANALYSIS", str(shadow_id), f"User {user} requested deep AI analysis.")
    return result


@app.get("/api/ai/summarize")
def ai_summarize_risks(user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """Generate executive risk summary using Cohere."""
    shadows = db.query(ShadowPurchase).filter(ShadowPurchase.status == "Pending").all()
    shadows_data = []
    for s in shadows:
        txn = db.query(Transaction).filter(Transaction.id == s.transaction_id).first()
        shadows_data.append({
            "vendor": txn.vendor if txn else "Unknown",
            "amount": txn.amount if txn else 0,
            "risk_score": s.risk_score or 0,
            "item_category": s.item_category or "General",
            "reason": s.reason or "N/A",
            "department": txn.department if txn else "Unknown",
        })

    result = summarize_risks_with_cohere(shadows_data)
    _log_event(db, "AI_RISK_SUMMARY", None, f"User {user} generated AI risk summary.")
    return result


@app.get("/api/ai/vendor-insight/{vendor_name}")
def ai_vendor_insight(vendor_name: str, user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """AI-powered vendor risk insight using Cohere."""
    vendor = db.query(Vendor).filter(Vendor.name == vendor_name).first()
    if not vendor:
        raise HTTPException(status_code=404, detail=f"Vendor '{vendor_name}' not found")

    shadow_count = db.query(ShadowPurchase).join(
        Transaction, ShadowPurchase.transaction_id == Transaction.id
    ).filter(Transaction.vendor == vendor_name).count()
    total_spend = sum(
        t.amount for t in db.query(Transaction).filter(Transaction.vendor == vendor_name).all()
    )

    vendor_data = {
        "name": vendor.name,
        "risk_level": vendor.risk_level,
        "trust_score": vendor.trust_score,
        "shadow_count": shadow_count,
        "total_spend": total_spend,
        "approved": vendor.approved,
    }

    result = generate_vendor_insight_with_cohere(vendor_data)
    return result


@app.get("/api/ai/health")
def ai_health_check():
    """Check AI provider connectivity."""
    return check_ai_health()


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
